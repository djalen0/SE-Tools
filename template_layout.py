import re

from layout_constants import (
    BOX_FIELD_SUFFIX_TO_KEY,
    METADATA_TOKEN_ORDER,
    METADATA_TOKEN_LABELS,
    METADATA_TOKEN_KEYS,
    BOX_FIELD_FORMULA_RE,
)

# --- Master spreadsheet layout --------------------------------------------
#
# design.xlsx is treated as a *live* template rather than a set of hardcoded
# row/column numbers: this code reads its own geometry straight out of the
# placeholder tokens already sitting in the file (things like "SEC1_BOX17"
# or a formula that builds "SEC1_BOX17_MODEL") so that if the layout, row
# spacing, or column widths in design.xlsx ever change, the script follows
# along automatically instead of drifting out of sync with a stale snapshot.
#
# The convention the design files use:
#   - design.xlsx defines a single section "card" (SEC1) -- a title, a box
#     list, and a metadata column. At runtime the user is prompted for how
#     many cards to place side by side in each row (2 per row, 3 per row,
#     etc.); however many sections come out of the parsed text file then get
#     tiled left-to-right in that grid, with additional rows of cards
#     stacking *downward* rather than the sheet growing wider and wider,
#     which is friendlier for printing.
#   - Each section's box list runs down 6 columns within its card (a row
#     label, then model / horizontal dispersion / angle / circuit / NFC),
#     with an adjacent metadata value column (aim, slider, weights, trims,
#     angles, picks).
#   - The box-field cells (model/dispersion/angle/circuit/NFC) are actually
#     formulas like =CONCATENATE($B5,"_MODEL") rather than literal text, so
#     detection below matches on the literal "_MODEL"-style fragment inside
#     the formula text itself -- this works whether or not the workbook was
#     last saved with cached calculated values.
#   - Only SEC1 is trusted for learning the geometry: if a template happens
#     to have more than one card pre-built, later ones have sometimes been
#     bulk-copy-pasted with mistakes (e.g. every row repeating "SEC3_BOX1"),
#     so SEC1 is treated as the one source of truth.

# Used only if design.xlsx is missing entirely, or its layout can't be
# recognized at all -- matches the latest known-good design.xlsx so a
# from-scratch workbook still looks right.
FALLBACK_LAYOUT = {
    'base_col': 4,
    'title_row': 3,
    # 1 more than before -- reserves both stripe gutter columns (hang/section
    # and circuit-set, base_col - 2 and base_col - 1 relative to it).
    'block_width': 11,
    'row_stride': 35,
    'box_header_row_offset': 1,
    'box_start_row_offset': 2,
    'max_boxes': 24,
    'col_offsets': {'label': 0, 'model': 1, 'dispersion': 2, 'angle': 3, 'circuit': 4, 'nfc': 5},
    'meta_col_offset': 7,
    'metadata_rows': [
        (METADATA_TOKEN_LABELS[t], off, METADATA_TOKEN_KEYS[t])
        for t, off in zip(METADATA_TOKEN_ORDER, [1, 4, 7, 10, 13, 16, 19, 22, 25, 28, 31])
    ],
    'header_labels': {0: 'Cab', 1: 'Model', 2: 'Horizontal Dispersion', 3: 'Angle', 4: 'Circuit #', 5: 'NFC'},
    # Keyed by column offset from base_col (0=label, 1-5=the 5 box fields,
    # 6=gap, 7=metadata) rather than a fixed-length list -- a real template
    # with fewer box fields selected (e.g. NFC deselected) has fewer offsets
    # than this, and scan_template_layout() below produces a dict shaped the
    # same way, sized to whatever that template actually has.
    'column_widths': {0: 6.25, 1: 9.04, 2: 8.62, 3: 7.93, 4: 4.87, 5: 7.23, 6: 4.25, 7: 10.88},
    'existing_section_count': 0,
    # None here means "no template font was captured" -- write_master_workbook
    # builds sensible generic Font objects lazily (Font isn't imported at
    # module scope, since openpyxl may not be installed at all).
    'title_font': None,
    'header_font': None,
    'box_field_fonts': {},
    'metadata_label_font': None,
    'metadata_value_font': None,
    'metadata_label_fill': None,
    'title_fill': None,
    'title_border': None,
    'header_border': None,
    'box_borders': {},
    'metadata_label_border': None,
    'metadata_value_border': None,
    'title_align': None,
    'header_align': None,
    'box_aligns': {},
    'metadata_label_align': None,
    'metadata_value_align': None,
    # Width of the two stripe gutters reserved at base_col - 2 (hang/section)
    # and base_col - 1 (circuit-set).
    'stripe_width': 3,
    'circuit_set_stripe_width': 3,
}


def _scan_placeholders(ws, max_row, max_col):
    """
    First pass over the worksheet: find every SEC<N>_TITLE, SEC1_BOX<N> (or
    SEC1_BOX<N>_NUM), and SEC1_<metadata token> placeholder cell. Returns
    (titles, box_rows, meta_cells, box_label_re), each dict keyed by the
    number/token found, valued (row, col) of the first cell seen for that
    key.
    """
    title_re = re.compile(r'^SEC(\d+)_TITLE$')
    box_label_re = re.compile(r'^SEC1_BOX(\d+)(?:_NUM)?$')  # design.xlsx has used both "SEC1_BOX1" and "SEC1_BOX1_NUM"
    meta_re = re.compile(r'^SEC1_(' + '|'.join(METADATA_TOKEN_ORDER) + r')$')

    titles = {}
    box_rows = {}
    meta_cells = {}

    for row_cells in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row_cells:
            v = cell.value
            if not isinstance(v, str):
                continue
            m = title_re.match(v)
            if m:
                titles.setdefault(int(m.group(1)), (cell.row, cell.column))
                continue
            m = box_label_re.match(v)
            if m:
                box_rows.setdefault(int(m.group(1)), (cell.row, cell.column))
                continue
            m = meta_re.match(v)
            if m:
                meta_cells.setdefault(m.group(1), (cell.row, cell.column))

    return titles, box_rows, meta_cells, box_label_re


def _count_box_capacity(ws, box_col, box_start_row, box_rows, box_label_re):
    """
    Box capacity: count contiguous matching rows starting at the first box
    row, rather than trusting the highest box *number* seen. A bulk-filled
    formula with a botched relative reference can repeat "BOX1, BOX2,
    BOX3, BOX1, BOX2, BOX3, ..." instead of incrementing all the way up --
    box_rows (keyed by that number, via setdefault) would then only ever
    see keys 1-3 and badly undercount a real 24-row template. Counting the
    physical run of box-label-shaped rows sidesteps the bad numbering
    entirely.
    """
    max_boxes = 0
    r = box_start_row
    while box_label_re.match(str(ws.cell(row=r, column=box_col).value or '')):
        max_boxes += 1
        r += 1
        if max_boxes > 500:
            break
    if max_boxes == 0:
        max_boxes = max(box_rows.keys())
    return max_boxes


def _scan_field_offsets(ws, base_col, box_rows, label_col_offset):
    """
    Field columns (model/dispersion/angle/circuit/nfc) are formulas, not
    literal placeholder text, so learn them by scanning a handful of box
    rows for a formula containing a literal "_MODEL"-style fragment. This
    scans the whole row rather than bounding by block width, since block
    width itself is derived below (partly *from* these offsets) -- the
    very specific "_MODEL"-style match, combined with left-to-right scan
    order and setdefault, means a second card further right (if one
    happens to exist) can never overwrite SEC1's own (first-seen) offsets.
    """
    col_offsets = {'label': label_col_offset}
    sample_rows = sorted(r for r, c in box_rows.values())[:6]
    for r in sample_rows:
        for cell in ws[r]:
            if cell.column < base_col:
                continue
            v = cell.value
            if not isinstance(v, str):
                continue
            m = BOX_FIELD_FORMULA_RE.search(v)
            if m:
                key = BOX_FIELD_SUFFIX_TO_KEY[m.group(1)]
                col_offsets.setdefault(key, cell.column - base_col)
    # Deliberately NOT force-filling every box field key with a fallback
    # offset here: a template that had a field (e.g. NFC) deselected in the
    # Fields tab genuinely has no "_NFC"-style formula anywhere in its box
    # rows, so col_offsets should end up *missing* that key entirely --
    # that's what tells worksheet_writer.py not to write/paint that column at
    # all, instead of writing cabinet data into whatever real column happens
    # to sit at that field's FALLBACK_LAYOUT offset (previously the gap
    # column, or worse, another card's territory).
    return col_offsets


def _scan_metadata_rows(ws, base_col, title_row, meta_cells):
    """
    Metadata rows (Aim/Slider/weights/trims/angles/picks), in the template's
    own declared order, using each token's own label text (one row above the
    value cell) so a cloned card matches the design file's exact wording.
    Falls back to FALLBACK_LAYOUT's rows entirely if the template has none
    at all.
    """
    metadata_rows = []
    meta_col_offset = None
    for token in METADATA_TOKEN_ORDER:
        if token in meta_cells:
            r, c = meta_cells[token]
            value_offset = r - title_row
            # Use the template's own label text (one row above the value
            # cell, same column) verbatim -- e.g. "Trim (T)" or "Aim" -- so
            # cloned cards match the design file's exact wording/casing
            # instead of a hardcoded guess.
            label_cell_value = ws.cell(row=r - 1, column=c).value
            label_text = label_cell_value if isinstance(label_cell_value, str) and label_cell_value else METADATA_TOKEN_LABELS[token]
            metadata_rows.append((label_text, value_offset - 1, METADATA_TOKEN_KEYS[token]))
            if meta_col_offset is None:
                meta_col_offset = c - base_col
    if not metadata_rows:
        metadata_rows = FALLBACK_LAYOUT['metadata_rows']
        meta_col_offset = FALLBACK_LAYOUT['meta_col_offset']
    return metadata_rows, meta_col_offset


def _derive_block_width(titles, base_col, col_offsets, meta_col_offset):
    """
    Horizontal stride between cards placed side by side. If the template
    happens to already have a second card built (SEC2_TITLE present), that
    tells us the gap the user actually wants, most authoritatively. If the
    template defines just one card (the normal "layout designer" case),
    derive a natural width from the card's own rightmost used column plus
    a one-column gap, instead of assuming a fixed stride that only matched
    one sample file.
    """
    if 2 in titles:
        return titles[2][1] - base_col
    # +4: one blank spacer column, then the next card's own two stripe
    # gutters (hang/section + circuit-set, base_col - 2 and base_col - 1
    # relative to it), then the next card itself.
    rightmost_offset = max(list(col_offsets.values()) + [meta_col_offset])
    return rightmost_offset + 4


def _derive_row_stride(titles, title_row, metadata_rows, box_start_row_offset, max_boxes):
    """
    Vertical stride between rows of cards. Same "trust an existing second
    row if the template already has one, otherwise infer a natural stride"
    approach as _derive_block_width.
    """
    if 3 in titles:
        return titles[3][0] - title_row
    # Only one row of cards exists in the template so far -- infer the
    # natural stride from how far down content runs, plus a small gap,
    # rather than hardcoding a number that only matched one sample file.
    # The metadata column runs *beside* the box list (not below it), so
    # take whichever extends further down: the metadata block or the
    # box list itself. Ignoring the box list here would let a large
    # box_count overrun into the next row of cards when there are few
    # metadata fields to compare against.
    metadata_extent = max(off for _, off, _ in metadata_rows)
    box_list_extent = box_start_row_offset + max_boxes - 1
    return max(metadata_extent, box_list_extent) + 4


def _scan_header_labels_and_widths(ws, base_col, header_row, col_offsets, meta_col_offset):
    """
    Header labels and column widths, both keyed by offset-from-base_col --
    scanned only for offsets this template actually uses (label + whichever
    box fields were actually selected, via col_offsets, plus the gap and
    metadata columns), not a fixed range(6)/range(8). A template with fewer
    box fields (e.g. NFC deselected) simply has fewer entries in each,
    which is exactly what lets build_new_block() clone a card of the right
    width instead of bleeding into whatever sits past its real last column.
    """
    from openpyxl.utils import get_column_letter

    gap_offset = meta_col_offset - 1
    used_offsets = sorted(set(col_offsets.values()) | {gap_offset, meta_col_offset})

    header_labels = {}
    for offset in used_offsets:
        if offset in (gap_offset, meta_col_offset):
            continue  # gap/metadata columns don't have a "box header" label
        val = ws.cell(row=header_row, column=base_col + offset).value
        header_labels[offset] = val if isinstance(val, str) and val else FALLBACK_LAYOUT['header_labels'].get(offset, '')

    column_widths = {}
    for offset in used_offsets:
        letter = get_column_letter(base_col + offset)
        dim = ws.column_dimensions.get(letter)
        width = dim.width if dim and dim.width else None
        column_widths[offset] = width if width else FALLBACK_LAYOUT['column_widths'].get(offset, 10)

    return header_labels, column_widths


def _scan_stripe_widths(ws, base_col):
    """Width of the two stripe gutters (hang/section at base_col-2,
    circuit-set at base_col-1), if the template has columns there at all."""
    from openpyxl.utils import get_column_letter

    stripe_dim = ws.column_dimensions.get(get_column_letter(base_col - 2)) if base_col > 2 else None
    stripe_width = (stripe_dim.width if stripe_dim and stripe_dim.width else None) or FALLBACK_LAYOUT['stripe_width']

    circuit_set_stripe_dim = ws.column_dimensions.get(get_column_letter(base_col - 1)) if base_col > 1 else None
    circuit_set_stripe_width = (
        (circuit_set_stripe_dim.width if circuit_set_stripe_dim and circuit_set_stripe_dim.width else None)
        or FALLBACK_LAYOUT['circuit_set_stripe_width']
    )
    return stripe_width, circuit_set_stripe_width


def _scan_styles(ws, base_col, title_row, header_row, box_start_row, col_offsets, metadata_rows, meta_col_offset):
    """
    Capture the template's *actual* fonts (size, bold, italic, color -- not
    just the name), fills, borders, and horizontal alignment for every role
    (title, header, box fields, metadata label/value) so that writing data
    never silently shrinks/re-styles cells back to some generic default, and
    cloned cards reuse these same styles too instead of a hardcoded guess.
    Only captured if the template cell actually has a real (non-default)
    value set -- an unstyled cell shouldn't force a particular style choice
    that wasn't really there.
    """
    import copy as _copy

    def _font_or_none(row, col):
        f = ws.cell(row=row, column=col).font
        return _copy.copy(f) if f and f.size else None

    title_font = _font_or_none(title_row, base_col)
    header_font = _font_or_none(header_row, base_col + 1)

    box_field_fonts = {}
    for key, offset in col_offsets.items():
        box_field_fonts[key] = _font_or_none(box_start_row, base_col + offset)

    metadata_label_font = None
    metadata_value_font = None
    metadata_label_fill = None
    metadata_label_border = None
    metadata_value_border = None
    metadata_label_align = None
    metadata_value_align = None
    if metadata_rows and meta_col_offset is not None:
        _, first_label_offset, _ = metadata_rows[0]
        first_value_row = title_row + first_label_offset + 1
        first_label_row = first_value_row - 1
        metadata_value_font = _font_or_none(first_value_row, base_col + meta_col_offset)
        metadata_label_font = _font_or_none(first_label_row, base_col + meta_col_offset)
        # Capture the metadata label's own fill too -- e.g. a custom
        # highlight color the user painted onto "Total Weight" and friends,
        # which is very often different from the box header row's fill.
        # Cloned cards need to reuse *this* fill, not the header's.
        label_fill = ws.cell(row=first_label_row, column=base_col + meta_col_offset).fill
        fg = label_fill.fgColor if label_fill else None
        if fg is not None and getattr(fg, 'rgb', None) not in (None, '00000000'):
            metadata_label_fill = _copy.copy(label_fill)
        metadata_label_border = _border_or_none(ws, first_label_row, base_col + meta_col_offset)
        metadata_value_border = _border_or_none(ws, first_value_row, base_col + meta_col_offset)
        metadata_label_align = _alignment_or_none(ws, first_label_row, base_col + meta_col_offset)
        metadata_value_align = _alignment_or_none(ws, first_value_row, base_col + meta_col_offset)

    # Title fill -- e.g. a banner color the user painted onto the title cell
    # -- and borders on the title/header/box cells, captured the same
    # cautious way as fonts: only if the template cell actually has one,
    # so a template with no styling at all doesn't force one into existence.
    title_fill = None
    title_cell_fill = ws.cell(row=title_row, column=base_col).fill
    fg = title_cell_fill.fgColor if title_cell_fill else None
    if fg is not None and getattr(fg, 'rgb', None) not in (None, '00000000'):
        title_fill = _copy.copy(title_cell_fill)

    title_border = _border_or_none(ws, title_row, base_col)
    header_border = _border_or_none(ws, header_row, base_col + 1)
    box_borders = {key: _border_or_none(ws, box_start_row, base_col + offset) for key, offset in col_offsets.items()}

    # Horizontal alignment, captured the same cautious way -- only if the
    # template cell actually has one set, so an unstyled template doesn't
    # force a particular alignment choice that wasn't really there.
    title_align = _alignment_or_none(ws, title_row, base_col)
    header_align = _alignment_or_none(ws, header_row, base_col + 1)
    box_aligns = {key: _alignment_or_none(ws, box_start_row, base_col + offset) for key, offset in col_offsets.items()}

    return {
        'title_font': title_font,
        'header_font': header_font,
        'box_field_fonts': box_field_fonts,
        'metadata_label_font': metadata_label_font,
        'metadata_value_font': metadata_value_font,
        'metadata_label_fill': metadata_label_fill,
        'title_fill': title_fill,
        'title_border': title_border,
        'header_border': header_border,
        'box_borders': box_borders,
        'metadata_label_border': metadata_label_border,
        'metadata_value_border': metadata_value_border,
        'title_align': title_align,
        'header_align': header_align,
        'box_aligns': box_aligns,
        'metadata_label_align': metadata_label_align,
        'metadata_value_align': metadata_value_align,
    }


def scan_template_layout(ws):
    """
    Reverse-engineer design.xlsx's real geometry from its own SEC1_* cells,
    instead of assuming any particular row/column numbers. Returns None if
    no recognizable SEC1_TITLE/SEC1_BOX1 pair is found, so the caller can
    fall back to FALLBACK_LAYOUT.

    Broken into one scan/derive step per concern (placeholders, box
    capacity, field offsets, metadata rows, grid stride, header/column
    widths, stripe widths, styles) so each piece can be read and tested on
    its own -- this used to be one ~280-line function covering all of it in
    sequence with no internal boundaries.
    """
    max_row = min(ws.max_row or 1, 5000)
    max_col = min(ws.max_column or 1, 80)
    titles, box_rows, meta_cells, box_label_re = _scan_placeholders(ws, max_row, max_col)

    if 1 not in titles or not box_rows:
        return None

    title_row, base_col = titles[1]

    box_start_row_offset = min(r for r, c in box_rows.values()) - title_row
    box_header_row_offset = box_start_row_offset - 1
    label_col_offset = min(c for r, c in box_rows.values()) - base_col

    box_col = base_col + label_col_offset
    box_start_row = title_row + box_start_row_offset
    max_boxes = _count_box_capacity(ws, box_col, box_start_row, box_rows, box_label_re)

    col_offsets = _scan_field_offsets(ws, base_col, box_rows, label_col_offset)
    metadata_rows, meta_col_offset = _scan_metadata_rows(ws, base_col, title_row, meta_cells)

    block_width = _derive_block_width(titles, base_col, col_offsets, meta_col_offset)
    row_stride = _derive_row_stride(titles, title_row, metadata_rows, box_start_row_offset, max_boxes)

    header_row = title_row + box_header_row_offset
    header_labels, column_widths = _scan_header_labels_and_widths(ws, base_col, header_row, col_offsets, meta_col_offset)
    stripe_width, circuit_set_stripe_width = _scan_stripe_widths(ws, base_col)

    styles = _scan_styles(ws, base_col, title_row, header_row, box_start_row, col_offsets, metadata_rows, meta_col_offset)

    return {
        'base_col': base_col,
        'title_row': title_row,
        'block_width': block_width,
        'row_stride': row_stride,
        'box_header_row_offset': box_header_row_offset,
        'box_start_row_offset': box_start_row_offset,
        'max_boxes': max_boxes,
        'col_offsets': col_offsets,
        'meta_col_offset': meta_col_offset,
        'metadata_rows': metadata_rows,
        'header_labels': header_labels,
        'column_widths': column_widths,
        'existing_section_count': max(titles.keys()),
        **styles,
        'stripe_width': stripe_width,
        'circuit_set_stripe_width': circuit_set_stripe_width,
    }


def scan_page_header(ws):
    """
    Find the page-level title/field placeholders (PAGE_TITLE, PAGE_<KEY>)
    that design_generator.py writes above the card grid, plus each field's
    label text and the title's current merge extent -- so
    write_master_workbook() can fill in real values and re-widen the title
    to match however many cards actually end up side by side, without
    needing to know anything about page-header layout ahead of time.

    Returns {'title_cell': (row, col) or None, 'title_rows': (first, last)
    or None, 'fields': {key: {'row', 'col', 'label'}}}. A template generated
    before this feature existed simply has no PAGE_TITLE cell at all, so
    title_cell comes back None and the caller skips the feature entirely --
    a pure opt-in, zero impact on older presets.
    """
    page_re = re.compile(r'^PAGE_(.+)$')

    title_cell = None
    fields = {}

    max_row = min(ws.max_row or 1, 200)
    max_col = min(ws.max_column or 1, 80)
    for row_cells in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row_cells:
            v = cell.value
            if not isinstance(v, str):
                continue
            m = page_re.match(v)
            if not m:
                continue
            key = m.group(1)
            if key == 'TITLE':
                title_cell = (cell.row, cell.column)
            else:
                # The label lives one column to the left, same row -- the
                # same "stacked label:value pairs" convention
                # design_generator.py writes.
                label_val = ws.cell(row=cell.row, column=cell.column - 1).value
                label = label_val if isinstance(label_val, str) and label_val else key.title()
                fields[key.lower()] = {'row': cell.row, 'col': cell.column, 'label': label}

    title_rows = None
    if title_cell:
        title_rows = (title_cell[0], title_cell[0])
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= title_cell[0] <= rng.max_row and rng.min_col <= title_cell[1] <= rng.max_col:
                title_rows = (rng.min_row, rng.max_row)
                break

    return {'title_cell': title_cell, 'title_rows': title_rows, 'fields': fields}


def _border_or_none(ws, row, col):
    """Copy a cell's border only if it actually has a visible line on at
    least one side -- an empty/default Border object shouldn't count as
    "the template has a border", the same caution used for fonts/fills."""
    import copy as _copy
    b = ws.cell(row=row, column=col).border
    if b and any(side.style for side in (b.left, b.right, b.top, b.bottom)):
        return _copy.copy(b)
    return None


def _alignment_or_none(ws, row, col):
    """Copy a cell's alignment only if a horizontal alignment is actually
    set -- an unstyled cell's default (None/'general') shouldn't count as
    "the template wants left/center/right", the same caution used for
    fonts/fills/borders."""
    import copy as _copy
    a = ws.cell(row=row, column=col).alignment
    if a and a.horizontal:
        return _copy.copy(a)
    return None


def merge_and_border_title(ws, row, col_start, col_end, border=None):
    """
    (Re-)merge a title cell across [col_start, col_end] on `row`, and if
    `border` is given, draw it as a single outer rectangle around the merged
    range (left edge only on the leftmost cell, right edge only on the
    rightmost, top/bottom on every cell in between) rather than a full grid,
    since it's one visual banner, not a table.

    Safe to call on a cell that's already merged (e.g. re-running against an
    already-formatted template) -- any existing merge overlapping this row's
    range is undone first, since openpyxl raises if you merge over a merge.
    """
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and not (rng.max_col < col_start or rng.min_col > col_end):
            ws.unmerge_cells(str(rng))

    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

    if border:
        from openpyxl.styles import Border
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).border = Border(
                left=border.left if c == col_start else None,
                right=border.right if c == col_end else None,
                top=border.top,
                bottom=border.bottom,
            )


def build_new_block(ws, block_start_col, block_row, title, layout, fonts, header_fill=None):
    """
    Create a brand-new section block -- used once a file has more sections
    than the template already has room for. Column widths, header labels,
    metadata labels, and fonts are all cloned from the layout/fonts dicts
    (themselves read straight off the template), never from a hardcoded
    snapshot, so a cloned card looks identical to the template's own card.
    """
    from openpyxl.utils import get_column_letter

    for offset, width in layout['column_widths'].items():
        ws.column_dimensions[get_column_letter(block_start_col + offset)].width = width

    if block_start_col > 2:
        stripe_letter = get_column_letter(block_start_col - 2)
        ws.column_dimensions[stripe_letter].width = layout.get('stripe_width', 3)
    if block_start_col > 1:
        circuit_set_stripe_letter = get_column_letter(block_start_col - 1)
        ws.column_dimensions[circuit_set_stripe_letter].width = layout.get('circuit_set_stripe_width', 3)

    # Title -- merged across the card's full width (box fields + metadata
    # column), same as a freshly generated template, so a cloned card's
    # title isn't left dependent on empty neighbor cells to display fully.
    title_cell = ws.cell(row=block_row, column=block_start_col, value=title)
    title_cell.font = fonts['title_font']
    if fonts.get('title_align'):
        title_cell.alignment = fonts['title_align']
    if fonts.get('title_fill'):
        title_cell.fill = fonts['title_fill']
    merge_end_offset = max(list(layout['col_offsets'].values()) + [layout['meta_col_offset']])
    merge_and_border_title(ws, block_row, block_start_col, block_start_col + merge_end_offset, fonts.get('title_border'))

    header_row = block_row + layout['box_header_row_offset']
    header_border = fonts.get('header_border')
    header_align = fonts.get('header_align')
    for offset in layout['header_labels']:
        cell = ws.cell(row=header_row, column=block_start_col + offset, value=layout['header_labels'].get(offset, ''))
        cell.font = fonts['header_font']
        if header_border:
            cell.border = header_border
        if header_align:
            cell.alignment = header_align
        if header_fill:
            cell.fill = header_fill

    # Metadata labels get *their own* fill (e.g. a custom highlight the user
    # painted onto "Total Weight" and friends), not the box header's gray --
    # those two are frequently different colors in a real template, and
    # blindly reusing header_fill here was overwriting a cloned card's
    # metadata labels with the wrong color.
    meta_fill = fonts.get('metadata_label_fill')
    meta_label_border = fonts.get('metadata_label_border')
    meta_label_align = fonts.get('metadata_label_align')
    meta_col = block_start_col + layout['meta_col_offset']
    for label, label_offset, _ in layout['metadata_rows']:
        cell = ws.cell(row=block_row + label_offset, column=meta_col, value=label)
        cell.font = fonts['metadata_label_font']
        if meta_label_border:
            cell.border = meta_label_border
        if meta_label_align:
            cell.alignment = meta_label_align
        if meta_fill:
            cell.fill = meta_fill
