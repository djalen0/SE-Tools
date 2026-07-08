import json
import os
from pathlib import Path

from gui_utils import win_long_path
from template_layout import scan_template_layout, build_new_block, merge_and_border_title, scan_page_header, FALLBACK_LAYOUT

# Cells this far right in a box row are left uncolored even when circuit
# color-coding is on -- matches the reference sheet this feature was modeled
# on, where the auxiliary PAN column stayed gray while HI-D/HORNS/PINS got
# the color band. NFC is our closest equivalent to that auxiliary column.
CIRCUIT_COLOR_EXCLUDED_FIELDS = {'nfc'}

# How to pull each box field's value out of a parsed cabinet dict. Iterated
# over whatever keys layout['col_offsets'] actually contains for this
# template (see the box-row write loop below) rather than a hardcoded tuple,
# so a field the user deselected in the Fields tab -- and which therefore has
# no real column in this template at all -- is simply never looked up or
# written, instead of landing in whatever column a stale fallback offset
# happened to point at.
CABINET_FIELD_GETTERS = {
    'label': lambda c: c['position'],
    'model': lambda c: c['model'],
    'dispersion': lambda c: c['dispersion'],
    'angle': lambda c: c['splay'],
    'circuit': lambda c: c['ckt'],
    'nfc': lambda c: c.get('nfc', ''),
}


def load_circuit_color_config(design_path):
    """
    Circuit-number color-coding config lives in a sidecar JSON next to the
    chosen design preset (same stem, .colors.json) rather than inside the
    workbook itself -- there's no template cell to scan it back out of, since
    it depends on which Circuit # values actually show up in the parsed
    data, not anything visible in the template. Presets generated before
    this feature existed simply have no sidecar file, which just means
    "no circuit coloring" -- a pure opt-in, no change to existing output.
    """
    disabled = {
        'enabled': False, 'circuit_colors': [], 'cycle_length': 4, 'hang_colors': [],
        'circuit_set_enabled': False, 'circuit_set_colors': [],
        'show_row_fill': True, 'hid_bundle_size': 4, 'breakout_cable_name': 'Hi-D',
    }
    if not design_path:
        return disabled
    colors_path = Path(design_path).with_suffix('.colors.json')
    if not os.path.exists(win_long_path(colors_path)):
        return disabled
    try:
        with open(win_long_path(colors_path), 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (OSError, ValueError):
        return disabled
    circuit_colors = data.get('circuit_colors') or []
    # Missing key (sidecar written before cycle_length existed) falls back to
    # using the whole list, preserving that preset's prior behavior exactly.
    cycle_length = data.get('cycle_length') or len(circuit_colors) or 4
    return {
        'enabled': bool(data.get('enabled', True)),
        'circuit_colors': circuit_colors,
        'cycle_length': cycle_length,
        'hang_colors': data.get('hang_colors') or [],
        # Missing keys (sidecar written before this feature existed) default
        # to disabled -- a pure opt-in, same as every other coloring feature
        # here, so an older preset's output doesn't change at all.
        'circuit_set_enabled': bool(data.get('circuit_set_enabled', False)),
        'circuit_set_colors': data.get('circuit_set_colors') or [],
        # Breakout-cable numbering preferences (which brand/cable name, how
        # many circuits per breakout, whether the row is fully painted) are
        # true user preferences, not per-file data -- so they're carried
        # through here too, the same way the colors above are, instead of
        # silently resetting to their defaults every time a brand-new input
        # file (no matching saved per-file state) is opened.
        'show_row_fill': bool(data.get('show_row_fill', True)),
        'hid_bundle_size': data.get('hid_bundle_size') or 4,
        'breakout_cable_name': data.get('breakout_cable_name') or 'Hi-D',
    }


def _contrast_text_color(argb_hex):
    """
    Pick black or white text so it reads clearly against a given ARGB fill,
    using the standard relative-luminance threshold -- so a user picking,
    say, yellow for a circuit color still gets legible text without having
    to separately choose a text color for every fill.
    """
    hex6 = str(argb_hex)[-6:]
    try:
        r, g, b = (int(hex6[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return 'FF000000'
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return 'FF000000' if luminance > 0.6 else 'FFFFFFFF'


def _hang_stripe_fill(header, hang_colors):
    """First matching hang_colors entry (case-insensitive substring against
    the section header), or None if nothing matches / the match is explicitly
    set to "no stripe" (fill: null). This paints a standalone identity
    stripe column next to the card -- it doesn't touch the row-level
    Circuit # coloring, which keeps cycling normally on every card
    regardless of hang."""
    header_lower = (header or '').lower()
    for entry in hang_colors:
        match = str(entry.get('match', '')).lower()
        if match and match in header_lower:
            return entry.get('fill')
    return None


def _circuit_identity(cabinet):
    """
    The value to group a cabinet by for circuit-color purposes: the
    ORIGINAL (pre-Hi-D) circuit number if the web editor sent one along
    (cabinet['_normalCkt'], preserved from the browser's own Hi-D-numbering
    state), otherwise the plain 'ckt' value. Once Hi-D/breakout numbering is
    applied in the browser, every breakout cable's legs get relabeled back
    to the same few strings (e.g. every breakout shows "4,3,2,1"), so
    grouping by the displayed 'ckt' label alone would collapse every
    breakout on the whole hang into one indistinguishable color group
    instead of cycling normally -- using the stable original number instead
    keeps row-fill and circuit-set coloring correct in the exported sheet
    regardless of which numbering mode was showing when it was exported.
    """
    normal = cabinet.get('_normalCkt')
    return normal if normal else cabinet.get('ckt')


def _assign_circuit_colors(cabinets, palette):
    """Map each distinct Circuit # value to the next color in `palette`, in
    first-seen order, wrapping around if there are more distinct circuits
    than palette entries."""
    assignment = {}
    if not palette:
        return assignment
    for cabinet in cabinets:
        ckt = _circuit_identity(cabinet)
        if ckt not in assignment:
            assignment[ckt] = palette[len(assignment) % len(palette)]
    return assignment


def _assign_circuit_set_colors(cabinets, palette, cycle_length):
    """
    Group distinct Circuit # values (in first-seen order) into consecutive
    sets of `cycle_length`, and map every circuit in a set to that set's
    color from `palette` (wrapping back to the start of the palette after 8
    sets). This is a second, independent grouping from the row-level circuit
    color cycle -- e.g. with cycle_length=4, circuits 1-4 are "Set 1" and all
    share one set color, circuits 5-8 are "Set 2" and share the next, and so
    on, regardless of which of the 4 circuit-fill colors each individual row
    ends up with.
    """
    assignment = {}
    if not palette:
        return assignment
    cycle_length = max(1, cycle_length)
    seen_order = []
    for cabinet in cabinets:
        ckt = _circuit_identity(cabinet)
        if ckt not in assignment:
            seen_order.append(ckt)
            set_index = (len(seen_order) - 1) // cycle_length
            assignment[ckt] = palette[set_index % len(palette)]
    return assignment


def _colored_font(base_font, contrast_hex):
    """Clone a Font with just the color swapped, so a circuit-colored cell
    keeps whatever name/size/bold the template's own box font already has."""
    from openpyxl.styles import Font
    return Font(
        name=base_font.name, size=base_font.size,
        bold=base_font.bold, italic=base_font.italic,
        color=contrast_hex,
    )


def describe_page_header(design_path):
    """
    Peek at a design preset to see whether it defines a page-level title/
    field block, and what fields it has -- so txt_parse_v5.py can prompt for
    values (or skip prompting entirely, for an older preset predating this
    feature) before write_master_workbook() runs. Returns {'has_title': bool,
    'fields': [{'key', 'label'}, ...]}, fields in the order they appear down
    the sheet.
    """
    design_path = Path(design_path) if design_path else None
    if not design_path or not os.path.exists(win_long_path(design_path)):
        return {'has_title': False, 'fields': []}
    from openpyxl import load_workbook
    wb = load_workbook(win_long_path(design_path))
    ws = wb.active
    page_header = scan_page_header(ws)
    ordered_fields = sorted(page_header['fields'].items(), key=lambda kv: kv[1]['row'])
    fields = [{'key': key, 'label': info['label']} for key, info in ordered_fields]
    return {'has_title': page_header['title_cell'] is not None, 'fields': fields}


def write_master_workbook(sections, design_path, output_path, cards_per_row=2, page_header_values=None):
    """
    Plot parsed sections into a workbook mirroring the master design file's
    layout: design.xlsx defines a single section "card" (SEC1), and this
    tiles however many sections were parsed into a grid of `cards_per_row`
    cards per row, with additional rows of cards stacking downward rather
    than the sheet growing wider and wider. If design_path points to a real
    workbook, it's loaded, its geometry is learned from SEC1's own
    placeholder cells, and its placeholder cells are overwritten in place
    (preserving whatever formatting/branding is already there). Cards beyond
    what the template already has built get a brand-new one cloned from the
    template's own widths/labels/fill. Falls back to a plain workbook from
    scratch if no design file is available.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

    design_path = Path(design_path) if design_path else None
    used_template = bool(design_path and os.path.exists(win_long_path(design_path)))
    circuit_color_config = load_circuit_color_config(design_path)

    layout = None
    if used_template:
        wb = load_workbook(win_long_path(design_path))
        ws = wb.active
        layout = scan_template_layout(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

    warnings = []
    if not used_template:
        warnings.append(f"No design.xlsx found at {design_path}; built a plain workbook instead.")
    elif layout is None:
        warnings.append("Could not recognize design.xlsx's layout (no SEC1_TITLE/SEC1_BOX1 placeholders found); used a default layout instead.")

    if layout is None:
        layout = FALLBACK_LAYOUT

    # Fonts: use whatever the template itself actually has for each role,
    # falling back to a plain generic font only when nothing was capturable
    # (e.g. no template at all). This is what keeps writing data from
    # silently resetting font size/style back to some generic default.
    box_field_fonts = layout.get('box_field_fonts') or {}
    default_data_font = Font(name='Arial', size=12)
    # A generic thin border, used only where the template itself didn't have
    # one captured -- so a from-scratch run (no template at all) still comes
    # out with visible card/table edges instead of bare, borderless text.
    default_thin = Side(style='thin', color='FF000000')
    default_border = Border(left=default_thin, right=default_thin, top=default_thin, bottom=default_thin)
    # Thick outer frame drawn around each card's full bounding box, on top of
    # the thin per-cell grid above -- a single unmissable "this is one card"
    # edge, independent of whatever the template's own per-cell borders are.
    thick_card_border = Side(style='thick', color='FF000000')
    # Vertical centering is always applied, even with no template at all
    # (Excel's own default is bottom-aligned, not centered); horizontal
    # alignment is only forced if the template actually specified one for
    # that role, otherwise a cell's horizontal alignment is left at Excel's
    # own default ('general').
    default_align = Alignment(vertical='center')
    fonts = {
        'title_font': layout.get('title_font') or Font(name='Arial', size=18),
        'header_font': layout.get('header_font') or Font(name='Arial', size=12),
        'metadata_label_font': layout.get('metadata_label_font') or Font(name='Arial', size=12),
        # None here just means "no custom fill was captured" -- build_new_block
        # then simply leaves a cloned label's fill untouched rather than
        # forcing some fallback color that might not match the template at all.
        'metadata_label_fill': layout.get('metadata_label_fill'),
        'title_fill': layout.get('title_fill'),
        'title_border': layout.get('title_border') or default_border,
        'header_border': layout.get('header_border') or default_border,
        'metadata_label_border': layout.get('metadata_label_border') or default_border,
        'title_align': layout.get('title_align') or default_align,
        'header_align': layout.get('header_align') or default_align,
        'metadata_label_align': layout.get('metadata_label_align') or default_align,
    }
    metadata_value_font = layout.get('metadata_value_font') or Font(name='Arial', size=11)
    metadata_value_border = layout.get('metadata_value_border') or default_border
    metadata_value_align = layout.get('metadata_value_align') or default_align
    box_data_fonts = {
        key: (box_field_fonts.get(key) or default_data_font)
        for key in ('label', 'model', 'dispersion', 'angle', 'circuit', 'nfc')
    }
    box_field_borders = layout.get('box_borders') or {}
    box_data_borders = {
        key: (box_field_borders.get(key) or default_border)
        for key in ('label', 'model', 'dispersion', 'angle', 'circuit', 'nfc')
    }
    box_field_aligns = layout.get('box_aligns') or {}
    box_data_aligns = {
        key: (box_field_aligns.get(key) or default_align)
        for key in ('label', 'model', 'dispersion', 'angle', 'circuit', 'nfc')
    }
    alarm_font = Font(
        name=metadata_value_font.name or 'Arial',
        size=metadata_value_font.size,
        bold=True,
        color='FFCC0000',
    )

    # Borrow the template's own header-row fill (if any) for cloned blocks,
    # rather than hardcoding a specific gray.
    header_fill = None
    if used_template:
        import copy
        sample_cell = ws.cell(row=layout['title_row'] + layout['box_header_row_offset'], column=layout['base_col'] + 1)
        fg = sample_cell.fill.fgColor if sample_cell.fill else None
        if fg is not None and getattr(fg, 'rgb', None) not in (None, '00000000'):
            header_fill = copy.copy(sample_cell.fill)

    def set_cell(row, col, value, font=None, border=None, alignment=None, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        return cell

    existing_count = layout.get('existing_section_count', 0)
    max_boxes = layout['max_boxes']
    co = layout['col_offsets']

    cards_per_row = max(1, int(cards_per_row))

    # Page header: a big centered document title above the whole grid, plus
    # whatever smaller Venue/Date/etc. fields the preset defines, filled in
    # once (not per section). A template built before this feature existed
    # simply has no PAGE_TITLE cell -- scan_page_header() then comes back
    # empty and this is skipped entirely, so older presets are unaffected.
    #
    # The title's merge is deliberately left exactly as design_generator.py
    # built it (card-1 width) rather than widened to match cards_per_row --
    # the Venue/Date/etc. fields sit in their own columns just past the
    # title's right edge (see generate_design()), and widening the merge to
    # span every card would swallow those field cells underneath it, hiding
    # their values. Only the cell values are set here, no merge changes.
    page_header = scan_page_header(ws) if used_template else {'title_cell': None, 'title_rows': None, 'fields': {}}
    if page_header['title_cell'] and page_header_values is not None:
        title_row_ph, title_col_ph = page_header['title_cell']
        ws.cell(row=title_row_ph, column=title_col_ph, value=page_header_values.get('title', ''))

        for key, info in page_header['fields'].items():
            ws.cell(row=info['row'], column=info['col'], value=page_header_values.get(key, ''))

    for section in sections:
        idx = section.get('section_number', 1)
        row_index = (idx - 1) // cards_per_row
        col_position = (idx - 1) % cards_per_row
        block_start_col = layout['base_col'] + layout['block_width'] * col_position
        block_row = layout['title_row'] + layout['row_stride'] * row_index

        if idx > existing_count:
            # The template doesn't have this many section cards yet.
            build_new_block(ws, block_start_col, block_row, section['header'], layout, fonts, header_fill)
        else:
            # Block already exists in the template -- leave its header row /
            # metadata labels / column widths alone (whatever the template
            # already has). The title is the one exception: it's always
            # (re-)merged across the card's full width and given the same
            # fill/border every other card gets, so long section names don't
            # get visually cut off depending on what happens to sit in the
            # neighboring cells, and every card's banner looks consistent.
            set_cell(block_row, block_start_col, section['header'], font=fonts['title_font'], alignment=fonts.get('title_align'))
            if fonts.get('title_fill'):
                ws.cell(row=block_row, column=block_start_col).fill = fonts['title_fill']
            merge_end_offset = max(list(co.values()) + [layout['meta_col_offset']])
            merge_and_border_title(ws, block_row, block_start_col, block_start_col + merge_end_offset, fonts.get('title_border'))

        # Cabinets
        cabinets = section['cabinets']
        if len(cabinets) > max_boxes:
            warnings.append(
                f"{section['header']}: {len(cabinets)} boxes found, but the master "
                f"layout only has room for {max_boxes}; extra boxes were skipped."
            )

        # Circuit # color-coding: the color cycle is assigned to each
        # distinct Circuit # value in first-seen order, repeating every
        # `cycle_length` colors -- this always applies, on every card,
        # regardless of hang. NFC is left uncolored either way (see
        # CIRCUIT_COLOR_EXCLUDED_FIELDS).
        circuit_coloring_on = circuit_color_config['enabled']
        cycle_length = max(1, min(circuit_color_config['cycle_length'], len(circuit_color_config['circuit_colors']) or 1))
        active_palette = circuit_color_config['circuit_colors'][:cycle_length]
        circuit_fill_map = _assign_circuit_colors(cabinets, active_palette) if circuit_coloring_on else {}

        # Outer gutter (block_start_col - 2): hang/section identity stripe --
        # a solid color for this card's full height (title through the last
        # box slot), independent of the row-level circuit-color cycle above,
        # if this section's header matches a hang_colors entry -- e.g. every
        # "Side" card gets a white stripe next to it while its box rows still
        # cycle normally.
        section_fill_hex = _hang_stripe_fill(section['header'], circuit_color_config['hang_colors']) if circuit_coloring_on else None
        stripe_last_row = block_row + layout['box_start_row_offset'] + max_boxes - 1
        if section_fill_hex and block_start_col > 2:
            outer_col = block_start_col - 2
            section_fill = PatternFill('solid', fgColor=section_fill_hex)
            for r in range(block_row, stripe_last_row + 1):
                ws.cell(row=r, column=outer_col).fill = section_fill

        # Inner gutter (block_start_col - 1): circuit-SET identity stripe --
        # every group of `cycle_length` distinct circuits gets its own color
        # from circuit_set_colors, for the rows that actually have box data.
        # Rows above the first box (title/header) and any unused box slots
        # past this card's real count "wrap over" to the outer/section color
        # instead of being left blank -- so a card with fewer boxes than
        # max_boxes still reads as one continuous section-colored block
        # outside its real data (matches the reference sheet this was
        # modeled on).
        circuit_set_on = circuit_color_config['circuit_set_enabled']
        circuit_set_palette = circuit_color_config['circuit_set_colors']
        circuit_set_fill_map = (
            _assign_circuit_set_colors(cabinets, circuit_set_palette, cycle_length)
            if circuit_set_on else {}
        )
        if (circuit_set_on or section_fill_hex) and block_start_col > 1:
            inner_col = block_start_col - 1
            wrap_fill = PatternFill('solid', fgColor=section_fill_hex) if section_fill_hex else None
            box_start_row_for_stripe = block_row + layout['box_start_row_offset']
            if wrap_fill:
                for r in range(block_row, box_start_row_for_stripe):
                    ws.cell(row=r, column=inner_col).fill = wrap_fill
            for i in range(max_boxes):
                row = box_start_row_for_stripe + i
                set_hex = circuit_set_fill_map.get(_circuit_identity(cabinets[i])) if i < len(cabinets) else None
                if set_hex:
                    ws.cell(row=row, column=inner_col).fill = PatternFill('solid', fgColor=set_hex)
                elif wrap_fill:
                    ws.cell(row=row, column=inner_col).fill = wrap_fill

        # Splay/Angle describes the joint with the box *above* it, not a
        # property of the box on its own -- Excel can't literally merge it
        # across the shared row boundary for every consecutive pair (box 2's
        # cell and box 3's cell would both need to claim box 2's own row,
        # which a single cell can't belong to two merges at once), so
        # instead each value stays in its own box's row (same as every
        # other field) but is top-aligned, pulling it visually up against
        # the border with the box above -- and box 1 (the topmost/reference
        # box, which has no box above it to describe an angle to) is left
        # blank instead of showing the old "Frame" placeholder.
        angle_top_align = Alignment(
            horizontal=(box_data_aligns.get('angle').horizontal if 'angle' in box_data_aligns else None),
            vertical='top',
        )

        box_start_row = block_row + layout['box_start_row_offset']
        for i in range(max_boxes):
            row = box_start_row + i
            if i < len(cabinets):
                cabinet = cabinets[i]

                fill_hex = circuit_fill_map.get(_circuit_identity(cabinet))
                row_fill = PatternFill('solid', fgColor=fill_hex) if fill_hex else None
                contrast_hex = _contrast_text_color(fill_hex) if fill_hex else None

                def field_font(key):
                    if row_fill and key not in CIRCUIT_COLOR_EXCLUDED_FIELDS:
                        return _colored_font(box_data_fonts[key], contrast_hex)
                    return box_data_fonts[key]

                def field_fill(key):
                    return row_fill if key not in CIRCUIT_COLOR_EXCLUDED_FIELDS else None

                # Only write the fields this template actually has (`co`, aka
                # layout['col_offsets'], only contains keys that were really
                # found in the template -- a deselected field like NFC is
                # simply absent here). Writing unconditionally to a
                # FALLBACK-guessed offset for a missing key used to spill
                # cabinet data into whatever real column happened to sit
                # there (the gap column, or worse).
                for key in co:
                    if key == 'angle':
                        value = '' if i == 0 else cabinet['splay']
                        alignment = angle_top_align
                    else:
                        value = CABINET_FIELD_GETTERS[key](cabinet)
                        alignment = box_data_aligns[key]
                    set_cell(row, block_start_col + co[key], value, font=field_font(key), border=box_data_borders[key], alignment=alignment, fill=field_fill(key))
            else:
                # No data for this box position -- blank it out rather than
                # leaving behind a leftover "SEC1_BOX17_MODEL"-style
                # placeholder from the template. Left uncolored regardless
                # of circuit coloring, same as the reference sheet doesn't
                # color past its last real row. Only touches fields this
                # template actually has, same reasoning as above.
                for key in co:
                    set_cell(row, block_start_col + co[key], '', font=box_data_fonts[key], border=box_data_borders[key], alignment=box_data_aligns[key])

        # Metadata values (values in ALARM get flagged in red/bold)
        alerts = set(section.get('safety_alerts', []))
        metadata = section.get('metadata', {})
        meta_col = block_start_col + layout['meta_col_offset']
        for label, label_offset, key in layout['metadata_rows']:
            value_row = block_row + label_offset + 1
            if key in alerts:
                set_cell(value_row, meta_col, 'ALARM', font=alarm_font, border=metadata_value_border, alignment=metadata_value_align)
            else:
                set_cell(value_row, meta_col, metadata.get(key, ''), font=metadata_value_font, border=metadata_value_border, alignment=metadata_value_align)

        # Thick outer border around the whole card -- title through the
        # lowest row actually used (the box list or the metadata column,
        # whichever runs further down), label column through the metadata
        # column. Overlaid last so it sits on top of the thin per-cell grid
        # already drawn above, rather than fighting with it.
        metadata_bottom_row = (
            block_row + max(off for _, off, _ in layout['metadata_rows']) + 1
            if layout['metadata_rows'] else block_row
        )
        card_bottom_row = max(stripe_last_row, metadata_bottom_row)
        card_right_col = block_start_col + layout['meta_col_offset']
        _apply_outer_border(ws, block_row, card_bottom_row, block_start_col, card_right_col, thick_card_border)

    wb.save(output_path)
    return warnings


def _apply_outer_border(ws, top_row, bottom_row, left_col, right_col, thick_side):
    """
    Overlay a thick rectangle border around a card's full bounding box (title
    through its last used row, label column through the metadata column),
    on top of whatever thin grid lines are already there. Corners get both a
    thick top/bottom edge and a thick left/right edge; the thin lines between
    cells stay exactly as they were -- only the four outer edges change.
    """
    from openpyxl.styles import Border
    for r in range(top_row, bottom_row + 1):
        for c in range(left_col, right_col + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            cell.border = Border(
                left=thick_side if c == left_col else existing.left,
                right=thick_side if c == right_col else existing.right,
                top=thick_side if r == top_row else existing.top,
                bottom=thick_side if r == bottom_row else existing.bottom,
            )
